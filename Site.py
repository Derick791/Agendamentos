import streamlit as st
from openpyxl import Workbook, load_workbook
from pathlib import Path
import pandas as pd

# ============================
# Configurações
# ============================

ARQUIVO = Path("inscricoes.xlsx")
LIMITE_VAGAS = 4

DATAS_TREINAMENTO = {
    "B1 - Substituir Caçamba Recuperadora Tipo Ponte": {
        "ADM (09-16h)": [
            "2025-09-22", "2025-09-29", "2025-10-06",
            "2025-10-13", "2025-10-20", "2025-10-27"
        ],
        "Noite (19h-02h)": [
            "2025-09-23", "2025-09-30", "2025-10-07",
            "2025-10-14", "2025-10-21", "2025-10-28"
        ],
    },
    "B2 - Substituir Cavaletes de Impacto articulado e rolos na mesa de impacto": {
        "ADM (09-16h)": [
            "2025-09-25", "2025-10-02", "2025-10-09",
            "2025-10-16", "2025-10-23", "2025-10-30"
        ],
        "Noite (19h-02h)": [
            "2025-09-26", "2025-10-03", "2025-10-10",
            "2025-10-17", "2025-10-24", "2025-10-31"
        ],
    },
    "B3 - Regular Freios Eletromagnéticos Do Giro da Lança Da EP2091KS e RCs 2092KS": {
        "ADM (09-16h)": [
            "2025-09-23", "2025-09-30", "2025-10-07",
            "2025-10-14", "2025-10-21", "2025-10-28"
        ],
        "Noite (19h-02h)": [
            "2025-09-25", "2025-10-02", "2025-10-09",
            "2025-10-16", "2025-10-23", "2025-10-30"
        ],
    },
    "B4 - Substituir Atuador de Freio Vulkan SH13": {
        "ADM (09-16h)": [
            "2025-09-26", "2025-10-03", "2025-10-10",
            "2025-10-17", "2025-10-24", "2025-10-31"
        ],
        "Noite (19h-02h)": [
            "2025-09-24", "2025-10-01", "2025-10-08",
            "2025-10-15", "2025-10-22", "2025-10-29"
        ],
    },
    "B5 - Realizar Substituição De Chapas De Revestimentos Silos e Chutes": {
        "ADM (09-16h)": [
            "2025-09-24", "2025-10-01", "2025-10-08",
            "2025-10-15", "2025-10-22", "2025-10-29"
        ],
        "Noite (19h-02h)": [
            "2025-09-22", "2025-09-29", "2025-10-06",
            "2025-10-13", "2025-10-20", "2025-10-27"
        ],
    },
}

# ============================
# Funções
# ============================

def criar_planilha():
    if not ARQUIVO.exists():
        wb = Workbook()
        ws = wb.active
        ws.title = "Inscricoes"
        ws.append([
            "Empresa", "Nome", "Matrícula", "Equipe/Gerência",
            "Treinamento", "Data", "Horário", "Turno"
        ])
        wb.save(ARQUIVO)

def vagas_disponiveis(data, horario):
    wb = load_workbook(ARQUIVO)
    ws = wb["Inscricoes"]
    usados = sum(1 for row in ws.iter_rows(min_row=2, values_only=True)
                 if row[5] == data and row[6] == horario)
    return max(LIMITE_VAGAS - usados, 0)

def salvar_inscricao(empresa, nome, matricula, equipe, treinamento, data, horario, turno):
    wb = load_workbook(ARQUIVO)
    ws = wb["Inscricoes"]

    # Verificar duplicidade
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[1] == nome and row[4] == treinamento and row[5] == data:
            st.error(f"{nome} já está inscrito neste treinamento nesta data.")
            return False

    if vagas_disponiveis(data, horario) <= 0:
        st.error(f"As vagas para {data} ({horario}) já se esgotaram.")
        return False

    ws.append([empresa, nome, matricula, equipe, treinamento, data, horario, turno])
    wb.save(ARQUIVO)
    return True

def carregar_inscricoes():
    wb = load_workbook(ARQUIVO)
    ws = wb["Inscricoes"]
    dados = [row for row in ws.iter_rows(min_row=2, values_only=True)]
    return pd.DataFrame(dados, columns=[
        "Empresa", "Nome", "Matrícula", "Equipe/Gerência",
        "Treinamento", "Data", "Horário", "Turno"
    ])

# ============================
# App Streamlit
# ============================

st.title("📌 Formulário de Treinamentos")
criar_planilha()

empresa = st.selectbox("Empresa", ["Vale", "Parceira"])
nome = st.text_input("Nome completo")

matricula = ""
if empresa == "Vale":
    matricula = st.text_input("Matrícula (8 dígitos)")
    if matricula and (not matricula.isdigit() or len(matricula) != 8):
        st.warning("A matrícula deve ter exatamente 8 dígitos numéricos.")

# Gerência ou Parceira
if empresa == "Vale":
    equipe = st.selectbox("Gerência", ["Gerência de Pátio", "Gerência de Usina"])
else:
    equipe = st.selectbox("Parceira", ["Usimig", "Plagecon", "NDT"])

# Treinamento -> Horário -> Data
treinamento = st.selectbox("Treinamento", list(DATAS_TREINAMENTO.keys()))
horarios_disponiveis = list(DATAS_TREINAMENTO[treinamento].keys())
horario = st.selectbox("Horário", horarios_disponiveis)
datas_disponiveis = DATAS_TREINAMENTO[treinamento][horario]
data = st.selectbox("Data", datas_disponiveis)

turno = st.selectbox("Turno", ["Turno A", "Turno B", "Turno C", "Turno D"])

# Mostrar vagas
if data and horario:
    disponiveis = vagas_disponiveis(data, horario)
    st.info(f"💺 Vagas disponíveis para {data} ({horario}): {disponiveis}/{LIMITE_VAGAS}")

# Botão salvar
if st.button("Salvar inscrição"):
    if not (empresa and nome and equipe and treinamento and data and horario and turno):
        st.warning("Preencha todos os campos obrigatórios.")
    elif empresa == "Vale" and (not matricula or len(matricula) != 8):
        st.warning("Matrícula inválida para funcionários da Vale.")
    else:
        if salvar_inscricao(empresa, nome, matricula, equipe, treinamento, data, horario, turno):
            st.success("✅ Inscrição registrada com sucesso!")

# Resumo
st.markdown("---")
st.subheader("📊 Resumo para o instrutor")

df = carregar_inscricoes()

if df.empty:
    st.info("Nenhuma inscrição registrada até o momento.")
else:
    contagem = (
        df.groupby(["Treinamento", "Data", "Horário"])
        .size()
        .reset_index(name="Inscritos")
    )
    contagem["Vagas Restantes"] = LIMITE_VAGAS - contagem["Inscritos"]

    st.write("### 👥 Turmas e vagas")
    st.dataframe(contagem)

    st.write("### 📋 Lista completa de inscritos")
    st.dataframe(df.sort_values(["Treinamento", "Data", "Horário"]))