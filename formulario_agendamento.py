import streamlit as st
from openpyxl import Workbook, load_workbook
from pathlib import Path
import pandas as pd
import io

# ============================ 
# Configura√ß√µes
# ============================

ARQUIVO = Path("inscricoes.xlsx")
LIMITE_VAGAS = 4

DATAS_TREINAMENTO = {
    "B1 - Substituir Ca√ßamba Recuperadora Tipo Ponte": {
        "ADM (09-16h)": [
            "2025-09-22", "2025-09-29", "2025-10-06", "2025-10-13", "2025-10-20", "2025-10-27",
            "2025-11-03", "2025-11-10", "2025-11-17", "2025-11-24", "2025-12-01", "2025-12-08",
            "2025-12-15", "2025-12-22", "2025-12-29"
        ],
        "Noite (19h-02h)": [
            "2025-09-23", "2025-09-30", "2025-10-07", "2025-10-14", "2025-10-21", "2025-10-28",
            "2025-11-04", "2025-11-11", "2025-11-18", "2025-11-25", "2025-12-02", "2025-12-09",
            "2025-12-16", "2025-12-23", "2025-12-30"
        ],
    },
    "B2 - Substituir Cavaletes de Impacto articulado e rolos na mesa de impacto": {
        "ADM (09-16h)": [
            "2025-09-25", "2025-10-02", "2025-10-09", "2025-10-16", "2025-10-23", "2025-10-30",
            "2025-11-06", "2025-11-13", "2025-11-20", "2025-11-27", "2025-12-04", "2025-12-11",
            "2025-12-18", "2025-12-25"
        ],
        "Noite (19h-02h)": [
            "2025-09-26", "2025-10-03", "2025-10-10", "2025-10-17", "2025-10-24", "2025-10-31",
            "2025-11-07", "2025-11-14", "2025-11-21", "2025-11-28", "2025-12-05", "2025-12-12",
            "2025-12-19", "2025-12-26"
        ],
    },
    "B3 - Regular Freios Eletromagn√©ticos Do Giro da Lan√ßa Da EP2091KS e RCs 2092KS": {
        "ADM (09-16h)": [
            "2025-09-23", "2025-09-30", "2025-10-07", "2025-10-14", "2025-10-21", "2025-10-28",
            "2025-11-04", "2025-11-11", "2025-11-18", "2025-11-25", "2025-12-02", "2025-12-09",
            "2025-12-16", "2025-12-23", "2025-12-30"
        ],
        "Noite (19h-02h)": [
            "2025-09-25", "2025-10-02", "2025-10-09", "2025-10-16", "2025-10-23", "2025-10-30",
            "2025-11-06", "2025-11-13", "2025-11-20", "2025-11-27", "2025-12-04", "2025-12-11",
            "2025-12-18", "2025-12-25"
        ],
    },
    "B4 - Substituir Atuador de Freio Vulkan SH13": {
        "ADM (09-16h)": [
            "2025-09-26", "2025-10-03", "2025-10-10", "2025-10-17", "2025-10-24", "2025-10-31",
            "2025-11-07", "2025-11-14", "2025-11-21", "2025-11-28", "2025-12-05", "2025-12-12",
            "2025-12-19", "2025-12-26"
        ],
        "Noite (19h-02h)": [
            "2025-09-24", "2025-10-01", "2025-10-08", "2025-10-15", "2025-10-22", "2025-10-29",
            "2025-11-05", "2025-11-12", "2025-11-19", "2025-11-26", "2025-12-03", "2025-12-10",
            "2025-12-17", "2025-12-24", "2025-12-31"
        ],
    },
    "B5 - Realizar Substitui√ß√£o De Chapas De Revestimentos Silos e Chutes": {
        "ADM (09-16h)": [
            "2025-09-24", "2025-10-01", "2025-10-08", "2025-10-15", "2025-10-22", "2025-10-29",
            "2025-11-05", "2025-11-12", "2025-11-19", "2025-11-26", "2025-12-03", "2025-12-10",
            "2025-12-17", "2025-12-24", "2025-12-31"
        ],
        "Noite (19h-02h)": [
            "2025-09-22", "2025-09-29", "2025-10-06", "2025-10-13", "2025-10-20", "2025-10-27",
            "2025-11-03", "2025-11-10", "2025-11-17", "2025-11-24", "2025-12-01", "2025-12-08",
            "2025-12-15", "2025-12-22", "2025-12-29"
        ],
    },
}
# ============================
# Fun√ß√µes
# ============================

def criar_planilha():
    if not ARQUIVO.exists():
        wb = Workbook()
        ws = wb.active
        ws.title = "Inscricoes"
        ws.append([
            "Empresa", "Nome", "Matr√≠cula", "Equipe/Ger√™ncia",
            "Treinamento", "Data", "Hor√°rio", "Turno"
        ])
        wb.save(ARQUIVO)

def vagas_disponiveis(data, horario):
    wb = load_workbook(ARQUIVO)
    ws = wb["Inscricoes"]
    usados = sum(1 for row in ws.iter_rows(min_row=2, values_only=True)
                 if row[5] == data and row[6] == horario)
    return max(LIMITE_VAGAS - usados, 0)

def salvar_inscricao(empresa, nome, matricula, equipe, treinamento, data, horario, turno):
    wb = load_workbook(ARQUIVO)
    ws = wb["Inscricoes"]

    # Verificar duplicidade
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[1] == nome and row[4] == treinamento and row[5] == data:
            st.error(f"{nome} j√° est√° inscrito neste treinamento nesta data.")
            return False

    if vagas_disponiveis(data, horario) <= 0:
        st.error(f"As vagas para {data} ({horario}) j√° se esgotaram.")
        return False

    ws.append([empresa, nome, matricula, equipe, treinamento, data, horario, turno])
    wb.save(ARQUIVO)
    return True

def carregar_inscricoes():
    wb = load_workbook(ARQUIVO)
    ws = wb["Inscricoes"]
    dados = [row for row in ws.iter_rows(min_row=2, values_only=True)]
    return pd.DataFrame(dados, columns=[
        "Empresa", "Nome", "Matr√≠cula", "Equipe/Ger√™ncia",
        "Treinamento", "Data", "Hor√°rio", "Turno"
    ])

# ============================
# App Streamlit
# ============================

st.title("üìå Formul√°rio de Treinamentos")
criar_planilha()

empresa = st.selectbox("Empresa", ["Vale", "Parceira"])
nome = st.text_input("Nome completo")

matricula = ""
if empresa == "Vale":
    matricula = st.text_input("Matr√≠cula (8 d√≠gitos)")
    if matricula and (not matricula.isdigit() or len(matricula) != 8):
        st.warning("A matr√≠cula deve ter exatamente 8 d√≠gitos num√©ricos.")

# Ger√™ncia ou Parceira
if empresa == "Vale":
    equipe = st.selectbox("Ger√™ncia", ["Ger√™ncia de P√°tio", "Ger√™ncia de Usina"])
else:
    equipe = st.selectbox("Parceira", ["Usimig", "Plagecon", "NDT"])

# Treinamento -> Hor√°rio -> Data
treinamento = st.selectbox("Treinamento", list(DATAS_TREINAMENTO.keys()))
horarios_disponiveis = list(DATAS_TREINAMENTO[treinamento].keys())
horario = st.selectbox("Hor√°rio", horarios_disponiveis)
datas_disponiveis = DATAS_TREINAMENTO[treinamento][horario]
data = st.selectbox("Data", datas_disponiveis)

turno = st.selectbox("Turno", ["Turno A", "Turno B", "Turno C", "Turno D"])

# Mostrar vagas
if data and horario:
    disponiveis = vagas_disponiveis(data, horario)
    st.info(f"üí∫ Vagas dispon√≠veis para {data} ({horario}): {disponiveis}/{LIMITE_VAGAS}")

# Bot√£o salvar
if st.button("Salvar inscri√ß√£o"):
    if not (empresa and nome and equipe and treinamento and data and horario and turno):
        st.warning("Preencha todos os campos obrigat√≥rios.")
    elif empresa == "Vale" and (not matricula or len(matricula) != 8):
        st.warning("Matr√≠cula inv√°lida para funcion√°rios da Vale.")
    else:
        if salvar_inscricao(empresa, nome, matricula, equipe, treinamento, data, horario, turno):
            st.success("‚úÖ Inscri√ß√£o registrada com sucesso!")

# Resumo
st.markdown("---")
st.subheader("üìä Resumo para o instrutor")

df = carregar_inscricoes()

if df.empty:
    st.info("Nenhuma inscri√ß√£o registrada at√© o momento.")
else:
    contagem = (
        df.groupby(["Treinamento", "Data", "Hor√°rio"])
        .size()
        .reset_index(name="Inscritos")
    )
    contagem["Vagas Restantes"] = LIMITE_VAGAS - contagem["Inscritos"]

    st.write("### üë• Turmas e vagas")
    st.dataframe(contagem)

    st.write("### üìã Lista completa de inscritos")
    st.dataframe(df.sort_values(["Treinamento", "Data", "Hor√°rio"]))
    
    st.markdown("---")
st.subheader("üì• Baixar banco de dados")

# L√™ o arquivo Excel atual
with open(ARQUIVO, "rb") as f:
    data = f.read()

st.download_button(
    label="‚¨áÔ∏è Baixar inscri√ß√µes (Excel)",
    data=data,
    file_name="inscricoes.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)